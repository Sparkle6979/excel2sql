import openpyxl
class ExcelObj(object):
    def __init__(self,filepath):
        self.filepath = filepath
        self.excel_wb = openpyxl.load_workbook(filename=filepath)
        self.table_num = len(self.excel_wb.sheetnames)

    def table2sql(self,table_name):
        # read the table from table_name
        tableobj = self.excel_wb[table_name]

        # get the information from the table by row
        datas = list(tableobj.iter_rows(values_only=True))

        # get the attributes from the first row
        table_attribute = list(datas[0])
        table_attribute.append('')

        lstmp,table_sql= list(),list()

        # create table sql,all the attributes type is varchar(255),and the name of the table is same as the sheet name
        separator = ' varchar(255),'
        createsql = 'create table {0} '.format(table_name) + '(' + separator.join(table_attribute)[:-1] + ');'
        table_sql.append(createsql)

        # insert information sql
        for i in range(1,len(datas)):
            for dataval in datas[i]:
                if dataval == None: lstmp.append('null')
                else:               lstmp.append("'" + str(dataval) + "'")          
            insertstr = 'insert into {0} values ('.format(table_name) + ','.join(lstmp) + ');'         
            lstmp.clear()
            table_sql.append(insertstr)

        # return the list of sql statement
        return table_sql

        

    def excel2sql(self):
        excel_sql = dict()
        for i in range(0,self.table_num):
            excel_sql[self.excel_wb.sheetnames[i]] = self.table2sql(self.excel_wb.sheetnames[i])

        # key: sheet name     values: sql statement
        return excel_sql